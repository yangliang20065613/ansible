#!/usr/bin/python

import sys
import os
import commands
import shutil
import signal
import xlsxwriter

from openpyxl import Workbook, load_workbook

bs = ['512B', '4k', '8k', '16k', '32k', '64k', '128k', '512k']
qd = ['1q', '2q', '4q', '8q', '16q', '32q', '64q', '128q'] 

def main():
    
    wb = Workbook()
    ws1 = wb.active
    ws1.title='2cpu'

    r_iops = 1
    r_bw = 9
    r_lat = 17
    c = 1

    for item_qd in qd:

        for item_bs in bs:
            file_name = item_bs + '-' + item_qd + '-' + "randread-2cpu"
            with open(file_name, 'r') as f:
                lines = f.readlines()
                last_line = lines[-2]
                result=last_line.split()

                ws1.cell(row=r_iops,column=c).value = result[2]
                ws1.cell(row=r_bw,column=c).value = result[3]
                ws1.cell(row=r_lat,column=c).value = result[4]
                r_iops = r_iops + 1
                r_bw = r_bw + 1
                r_lat = r_lat + 1
                print(result)
        c = c + 1
        r_iops = 1
        r_bw = 9
        r_lat = 17
         
    wb.save('2cpu.xlsx')
if __name__ == '__main__':
    main()
