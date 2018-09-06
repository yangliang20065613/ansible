#!/usr/bin/python

import sys
import os
import commands
import shutil
import signal

from openpyxl import Workbook, load_workbook

bs = ['512B', '4k', '8k', '16k', '32k', '64k', '128k', '512k']
qd = ['1q', '2q', '4q', '8q', '16q', '32q', '64q', '128q'] 

def main():
    
    wb_randread = Workbook()
    ws1_randread = wb_randread.active
    ws1_randread.title='2cpu-randread'

    r_iops = 1
    r_bw = 9
    r_lat = 17
    c = 1

    for item_qd in qd:

        for item_bs in bs:
            file_name = item_bs + '-' + item_qd + '-' + "randread-2cpu"
            with open(file_name, 'r') as f:
                lines = f.readlines()
                last_line = lines[-2]
                result=last_line.split()

                ws1_randread.cell(row=r_iops,column=c).value = result[2]
                ws1_randread.cell(row=r_bw,column=c).value = result[3]
                ws1_randread.cell(row=r_lat,column=c).value = result[4]
                r_iops = r_iops + 1
                r_bw = r_bw + 1
                r_lat = r_lat + 1
        c = c + 1
        r_iops = 1
        r_bw = 9
        r_lat = 17
    c = 1    
     
    wb_randread.save('2cpu-randread.xlsx')

    wb_randwrite = Workbook()
    ws1_randwrite = wb_randwrite.active
    ws1_randwrite.title='2cpu-randwrite'

    for item_qd in qd:

        for item_bs in bs:
            file_name = item_bs + '-' + item_qd + '-' + "randwrite-2cpu"
            with open(file_name, 'r') as f:
                lines = f.readlines()
                last_line = lines[-2]
                result=last_line.split()

                ws1_randwrite.cell(row=r_iops,column=c).value = result[2]
                ws1_randwrite.cell(row=r_bw,column=c).value = result[3]
                ws1_randwrite.cell(row=r_lat,column=c).value = result[4]
                r_iops = r_iops + 1
                r_bw = r_bw + 1
                r_lat = r_lat + 1
        c = c + 1
        r_iops = 1
        r_bw = 9
        r_lat = 17
    c = 1         

    wb_randwrite.save('2cpu-randwrite.xlsx')


    wb_read = Workbook()
    ws1_read = wb_read.active
    ws1_read.title='2cpu-read'

    for item_qd in qd:

        for item_bs in bs:
            file_name = item_bs + '-' + item_qd + '-' + "read-2cpu"
            with open(file_name, 'r') as f:
                lines = f.readlines()
                last_line = lines[-2]
                result=last_line.split()

                ws1_read.cell(row=r_iops,column=c).value = result[2]
                ws1_read.cell(row=r_bw,column=c).value = result[3]
                ws1_read.cell(row=r_lat,column=c).value = result[4]
                r_iops = r_iops + 1
                r_bw = r_bw + 1
                r_lat = r_lat + 1
        c = c + 1
        r_iops = 1
        r_bw = 9
        r_lat = 17
    c = 1    
    wb_read.save('2cpu-read.xlsx')

    wb_write = Workbook()
    ws1_write = wb_write.active
    ws1_write.title='2cpu-write'

    for item_qd in qd:

        for item_bs in bs:
            file_name = item_bs + '-' + item_qd + '-' + "write-2cpu"
            with open(file_name, 'r') as f:
                lines = f.readlines()
                last_line = lines[-2]
                result=last_line.split()

                ws1_write.cell(row=r_iops,column=c).value = result[2]
                ws1_write.cell(row=r_bw,column=c).value = result[3]
                ws1_write.cell(row=r_lat,column=c).value = result[4]
                r_iops = r_iops + 1
                r_bw = r_bw + 1
                r_lat = r_lat + 1
        c = c + 1
        r_iops = 1
        r_bw = 9
        r_lat = 17
    c = 1     
    wb_write.save('2cpu-write.xlsx')

if __name__ == '__main__':
    main()
