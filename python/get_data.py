#!/usr/bin/python

import sys
import os
import commands
import shutil
import signal
import xlsxwriter

from openpyxl import Workbook, load_workbook

bs = ['512B', '4k', '8k', '16k', '32k', '64k', '128k', '512k']
qd = ['1q', '2q', '4q', '8q', '16q', '32q', '64q', '128q'] 

def main():
    

    
    wb = Workbook()
    ws1 = wb.active
    ws1.title='2cpu'

    r = 1
    c = 1

    for item_qd in qd:

        ws1.cell(row=r,column=c).value = item_qd
        r = r + 1
 
        ws1.cell(row=r,column=c).value = ''
        ws1.cell(row=r,column=c+1).value = ''
        ws1.cell(row=r,column=c+2).value = ''
        r = r + 1

        for item_bs in bs:
            file_name = item_bs + '-' + item_qd + '-' + "randread-2cpu"
            with open(file_name, 'r') as f:
                lines = f.readlines()
                last_line = lines[-2]
                result=last_line.split()
                ws1.cell(row=r,column=c).value = result[2]
                ws1.cell(row=r,column=c+1).value = result[3]
                ws1.cell(row=r,column=c+2).value = result[4]
                r = r + 1
                print(result)
         
        ws1.cell(row=r,column=c).value = ''
        ws1.cell(row=r,column=c+1).value = ''
        ws1.cell(row=r,column=c+2).value = ''
        r = r + 1

    wb.save('2cpu.xlsx')
if __name__ == '__main__':
    main()
