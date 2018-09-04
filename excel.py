#!/usr/bin/python

import sys
import os
import commands
import shutil
import signal


from openpyxl import Workbook, load_workbook
wb = Workbook()
ws1 = wb.active
ws1.title='test'
ws1.cell(row=1,column=1).value = 100
ws1.cell(row=1,column=2).value = 180
ws1.cell(row=1,column=3).value = 150
wb.save('test.xlsx')
exit()

