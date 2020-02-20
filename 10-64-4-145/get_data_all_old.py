#!/usr/bin/python

import sys
import os
import commands
import shutil
import signal

from openpyxl import Workbook, load_workbook

bs = ['512', '1024', '2048', '4096', '8192', '16384', '32768', '65536', '131072', '524288']
qd = ['1', '2', '4', '8', '16', '32', '64', '128']

def main():
    
    wb_randread = Workbook()
    ws1_randread = wb_randread.active
    ws1_randread.title='1cpu-1cpu-randread'

    r_iops = 1
    r_bw = 11
    r_lat = 21
    c = 1

    for item_qd in qd:

        for item_bs in bs:
            file_name = item_bs + '-' + item_qd + '-' + "randread-1cpu-1pcu"
            with open(file_name, 'r') as f:
                lines = f.readlines()
                last_line = lines[-2]
                result=last_line.split()

                ws1_randread.cell(row=r_iops,column=c).value = result[2]
                ws1_randread.cell(row=r_bw,column=c).value = result[3]
                ws1_randread.cell(row=r_lat,column=c).value = result[4]
                r_iops = r_iops + 1
                r_bw = r_bw + 1
                r_lat = r_lat + 1
        c = c + 1
        r_iops = 1
        r_bw = 11
        r_lat = 21
    c = 1    
     
    wb_randread.save('1cpu-1pcu-randread.xlsx')

    wb_randwrite = Workbook()
    ws1_randwrite = wb_randwrite.active
    ws1_randwrite.title='1cpu-1pcu-randwrite'

    for item_qd in qd:

        for item_bs in bs:
            file_name = item_bs + '-' + item_qd + '-' + "randwrite-1cpu-1pcu"
            with open(file_name, 'r') as f:
                lines = f.readlines()
                last_line = lines[-2]
                result=last_line.split()

                ws1_randwrite.cell(row=r_iops,column=c).value = result[2]
                ws1_randwrite.cell(row=r_bw,column=c).value = result[3]
                ws1_randwrite.cell(row=r_lat,column=c).value = result[4]
                r_iops = r_iops + 1
                r_bw = r_bw + 1
                r_lat = r_lat + 1
        c = c + 1
        r_iops = 1
        r_bw = 11
        r_lat = 21
    c = 1         

    wb_randwrite.save('1cpu-1pcu-randwrite.xlsx')


    wb_read = Workbook()
    ws1_read = wb_read.active
    ws1_read.title='1cpu-1pcu-read'

    for item_qd in qd:

        for item_bs in bs:
            file_name = item_bs + '-' + item_qd + '-' + "read-1cpu-1pcu"
            with open(file_name, 'r') as f:
                lines = f.readlines()
                last_line = lines[-2]
                result=last_line.split()

                ws1_read.cell(row=r_iops,column=c).value = result[2]
                ws1_read.cell(row=r_bw,column=c).value = result[3]
                ws1_read.cell(row=r_lat,column=c).value = result[4]
                r_iops = r_iops + 1
                r_bw = r_bw + 1
                r_lat = r_lat + 1
        c = c + 1
        r_iops = 1
        r_bw = 11
        r_lat = 21
    c = 1    
    wb_read.save('1cpu-1pcu-read.xlsx')

    wb_write = Workbook()
    ws1_write = wb_write.active
    ws1_write.title='1cpu-1pcu-write'

    for item_qd in qd:

        for item_bs in bs:
            file_name = item_bs + '-' + item_qd + '-' + "write-1cpu-1pcu"
            with open(file_name, 'r') as f:
                lines = f.readlines()
                last_line = lines[-2]
                result=last_line.split()

                ws1_write.cell(row=r_iops,column=c).value = result[2]
                ws1_write.cell(row=r_bw,column=c).value = result[3]
                ws1_write.cell(row=r_lat,column=c).value = result[4]
                r_iops = r_iops + 1
                r_bw = r_bw + 1
                r_lat = r_lat + 1
        c = c + 1
        r_iops = 1
        r_bw = 11
        r_lat = 21
    c = 1     
    wb_write.save('1cpu-1pcu-write.xlsx')

if __name__ == '__main__':
    main()
