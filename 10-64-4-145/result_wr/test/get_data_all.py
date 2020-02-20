#!/usr/bin/python

import sys
import os
import commands
import shutil
import signal

from openpyxl import Workbook, load_workbook

bs = ['512b', '1k', '2k', '4k', '8k', '16k', '32k', '64k', '128k', '512k']
qd = ['1', '2', '4', '8', '16', '32', '64', '128']
#rw = ['write', 'read', 'randwrite', 'randread']

rw = ['write', 'read']

def main():

    for item_rw in rw:
    
        wb = Workbook()
        ws1 = wb.active
        ws1.title='1cpu-1cpu-' + item_rw 
    
        r_iops = 1
        r_bw = 11
        r_lat = 21
        c = 1
    
        for item_qd in qd:
    
            for item_bs in bs:
                file_name = item_bs + '-' + item_qd + 'q-' + item_rw + "-1cpu-1cpu"
                with open(file_name, 'r') as f:
                    lines = f.readlines()
                    result_iops_line = lines[13]
                    result_iops=result_iops_line.split()
                    iops = result_iops[1]
                    iops = iops[5:]
                    if iops[-2] == 'k':
                        iops = iops[:-2]
                    else:
                        iops = iops[:-1]
                        iops = iops[:-3] + "." + iops[-3:] 
    
#                    print(iops)
    
                    result_bw_line = lines[-1]
                    result_bw=result_bw_line.split()
                    bw = result_bw[2]
                    bw = bw[1:]
                    bw = bw[:-6]
#                    print(bw)
    
                    result_lat_line = lines[16]
                    result_lat = result_lat_line.split()
                    lat = result_lat[4]
                    lat = lat[4:]
                    lat = lat[:-1]
#                    print(lat)
    
                    ws1.cell(row=r_iops,column=c).value = iops
                    ws1.cell(row=r_bw,column=c).value = bw
                    ws1.cell(row=r_lat,column=c).value = lat
                    r_iops = r_iops + 1
                    r_bw = r_bw + 1
                    r_lat = r_lat + 1
            c = c + 1
            r_iops = 1
            r_bw = 11
            r_lat = 21
        
        result_file_name_xlsx = '1cpu-1pcu-' + item_rw + '.xlsx' 
        wb.save(result_file_name_xlsx)

if __name__ == '__main__':
    main()
